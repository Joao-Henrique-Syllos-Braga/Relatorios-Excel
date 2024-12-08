import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from main import data
from dado import datatotal


def createExcel(file, localFile):
    print("Começo da execução")
    
    # Verificar se o diretório existe
    if not os.path.exists(localFile):
        print(f"O diretório {localFile} não existe. Não foi possível salvar o arquivo.")
        return  # Retorna sem salvar o arquivo, pois o diretório não existe

    # Criar um novo arquivo Excel
    workbook = Workbook()

    # Adicionar as sheets para 2024 e 2023
    sheet_2024 = workbook.active
    sheet_2024.title = "2024"

    sheet_2023 = workbook.create_sheet(title="2023")

    sheet_total = workbook.create_sheet(title="Total")

    # Função para formatar números negativos e com duas casas decimais
    def format_number(value):
        if isinstance(value, (int, float)):
            # Formatar números negativos com sinal e limitar a 2 casas decimais
            return f"{value:,.2f}" if value >= 0 else f"- {abs(value):,.2f}"
        return value  # Retornar outros tipos de valor sem alteração

    # Função para aplicar formatação em cada sheet
    def format_sheet(sheet, dados):
        # Remover linhas de grade
        sheet.sheet_view.showGridLines = False

        # Definir largura da primeira coluna
        sheet.column_dimensions["A"].width = 30  # Aumentando a largura da primeira coluna

        # Ajustar a largura das outras colunas para o tamanho original
        for col in range(2, len(dados[0]) + 1):  # Ajustando as colunas a partir da segunda
            col_letter = chr(64 + col)  # Convertendo número da coluna para letra
            sheet.column_dimensions[col_letter].width = 10  # Definindo a largura original (ajuste conforme necessário)

        # Aumentar a largura da última coluna
        last_col = len(dados[0])  # Encontrar o número da última coluna
        last_col_letter = chr(64 + last_col)  # Converter para a letra correspondente da última coluna
        sheet.column_dimensions[last_col_letter].width = 25  # Ajuste a largura da última coluna como desejar

        # Preencher os dados na planilha
        for row_index, row in enumerate(dados, start=1):  # Iniciar o índice em 1 (Excel começa na linha 1)
            for col_index, value in enumerate(row, start=1):  # Iniciar o índice em 1 (Coluna A)
                # Substituir valores 0 por células vazias
                cell_value = "" if value == 0 else format_number(value)
                
                # Definir o valor na célula
                cell = sheet.cell(row=row_index, column=col_index, value=cell_value)
                
                # Alinhar os valores: alinhar à esquerda na primeira coluna e à direita nas outras
                if col_index == 1:  # Alinhar à esquerda na primeira coluna
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:  # Alinhar à direita nas outras colunas
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                
                # Definir o fundo branco para todas as linhas
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Branco

            # Ajustar a altura da linha (primeira linha mais alta, outras linhas mais baixas)
            if row_index == 1:
                sheet.row_dimensions[row_index].height = 20  # Altura mais moderada para a primeira linha
            else:
                sheet.row_dimensions[row_index].height = 18  # Altura mais baixa para as outras linhas

        # Definir a formatação para a primeira linha (cabeçalho)
        header_row = sheet[1]
        for cell in header_row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")  # Cor azul marinho
            cell.font = Font(color="FFFFFF")  # Texto branco

    # print("Obtenção de dados")
    # Obter os dados de 2024 e 2023
    # print("Obtenção de dados 01")
    dados_2024 = data(2024, file)
    # print("Obtenção de dados 02")
    dados_2023 = data(2023, file)
    # print("Obtenção de dados 2")
    dados_total = datatotal(file)


    # print("Formatar sheets")
    # Formatar as sheets
    format_sheet(sheet_2024, dados_2024)
    format_sheet(sheet_2023, dados_2023)
    format_sheet(sheet_total, dados_total)

    # print("Nome do arquivo")
    nome_arquivo = "00-Acompanhamento SAB TECH-OUT 24.xlsx"

    # print("Join OS")
    caminho_completo = os.path.join(localFile, nome_arquivo)
    print(caminho_completo)

    # Salvar o arquivo no caminho correto
    workbook.save(caminho_completo)
    print("Arquivo salvo no caminho:", caminho_completo)
